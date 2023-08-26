import os
import openpyxl
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM

def get_translation_direction():
    while True:
        print("Çeviri Yönünü Seçin:")
        print("1. İngilizce-Türkçe")
        print("2. Türkçe-İngilizce ")
        choice = input("Seçiminizi yapın (1 veya 2): ")
        if choice == '1':
            return "turkish"
        elif choice == '2':
            return "english"
        else:
            print("Geçersiz seçim. Lütfen tekrar deneyin.")

def get_sheet_selection(sheet_names):
    while True:
        print("Seçilebilecek Sayfalar:")
        print("0. Tüm Sayfaları Seç")
        for idx, sheet_name in enumerate(sheet_names, start=1):
            print(f"{idx}. {sheet_name}")
        try:
            choice = int(input("Lütfen bir sayfa seçin: "))
            if 0 <= choice <= len(sheet_names):
                return choice
            else:
                print("Geçersiz seçim. Lütfen tekrar deneyin.")
        except ValueError:
            print("Lütfen geçerli bir sayı girin.")

def get_cell_range(sheet_name):
    while True:
        range_choice = input(f"{sheet_name} için hücre aralığı belirtmek ister misiniz? (E/H): ")
        if range_choice.lower() == 'e':
            while True:
                range_str = input("Lütfen hücre aralığını girin (örn. A1:C10): ")
                try:
                    cell_range = workbook[sheet_name][range_str]
                    return cell_range
                except ValueError:
                    print("Geçersiz aralık. Lütfen tekrar deneyin.")
        elif range_choice.lower() == 'h':
            return workbook[sheet_name]['A1':f'{openpyxl.utils.get_column_letter(workbook[sheet_name].max_column)}{workbook[sheet_name].max_row}']
        else:
            print("Geçersiz seçim. Lütfen tekrar deneyin.")

def translate_text(text, tokenizer, model):
    cleaned_text = clean_special_tokens(text, tokenizer)
    translated_text = model.generate(**tokenizer(cleaned_text, return_tensors="pt", padding=True), max_new_tokens=100)
    translated_text = tokenizer.decode(translated_text[0], skip_special_tokens=True)
    return translated_text

def clean_special_tokens(text, tokenizer):
    text = text.replace(tokenizer.pad_token, "").replace(tokenizer.eos_token, "")
    return text.strip()

if __name__ == "__main__":
    cache_dir = os.getcwd()
    print("Yapay zeka destekli çeviri modelleri yükleniyor... lütfen bekleyin. (İnternet kullanımı olmaz.)")
    turkish_model_name = "Helsinki-NLP/opus-tatoeba-en-tr"
    turkish_tokenizer = AutoTokenizer.from_pretrained(turkish_model_name, cache_dir=cache_dir)
    turkish_model = AutoModelForSeq2SeqLM.from_pretrained(turkish_model_name, cache_dir=cache_dir)

    english_model_name = "Helsinki-NLP/opus-mt-tr-en"
    english_tokenizer = AutoTokenizer.from_pretrained(english_model_name, cache_dir=cache_dir)
    english_model = AutoModelForSeq2SeqLM.from_pretrained(english_model_name, cache_dir=cache_dir)

    current_directory = os.getcwd()
    excel_files = [file for file in os.listdir(current_directory) if file.endswith('.xlsx')]

    if not excel_files:
        print("Dizinde hiç Excel dosyası bulunamadı.")
    else:
        print("Seçilebilecek Excel Dosyaları:")
        for idx, file in enumerate(excel_files, start=1):
            print(f"{idx}. {file}")
        try:
            choice = int(input("Lütfen bir Excel dosyası seçin: "))
            if 1 <= choice <= len(excel_files):
                selected_file = excel_files[choice - 1]
                workbook = openpyxl.load_workbook(selected_file)

                translation_direction = get_translation_direction()
                sheet_names = workbook.sheetnames

                selected_sheet_index = get_sheet_selection(sheet_names)
                if selected_sheet_index == 0:
                    selected_sheets = sheet_names
                else:
                    selected_sheet = sheet_names[selected_sheet_index - 1]
                    selected_sheets = [selected_sheet]

                if translation_direction == "turkish":
                    tokenizer = turkish_tokenizer
                    model = turkish_model
                else:
                    tokenizer = english_tokenizer
                    model = english_model

                for sheet_name in selected_sheets:
                    cell_range = get_cell_range(sheet_name)

                    for row in cell_range:
                        for cell in row:
                            if cell.value:
                                translated_text = translate_text(cell.value, tokenizer, model)
                                cell.value = translated_text

                workbook.save(f"translated_{selected_file}")
                print(f"'{selected_file}' dosyası çevrildi ve 'translated_{selected_file}' olarak kaydedildi.")
                input("Push enter...")
            else:
                print("Geçersiz seçim. Lütfen tekrar deneyin.")
        except ValueError:
            print("Lütfen geçerli bir sayı girin.")
